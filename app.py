import pandas as pd
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import sys

# --- CONFIGURATION & INPUT ---
# Using sys.argv to take inputs from the Docker command line
course = sys.argv[1] if len(sys.argv) > 1 else "CS_101"
csv_file = sys.argv[2] if len(sys.argv) > 2 else "canvas_grades.csv"
DB_URL = "sqlite:///unt_data.db"
engine = create_engine(DB_URL)

def generate_excel_from_df(df, filename):
    """Generates a styled Excel report based on the provided DataFrame."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis Report"

    # Header row
    ws.append(list(df.columns))

    # Styles for Actionable UX (Red for at_risk, Green for high scorers)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for _, row in df.iterrows():
        ws.append(list(row))
        curr_row = ws.max_row
        
        # Apply conditional formatting based on 'status' or 'Total'
        if 'status' in df.columns and row['status'] == 'at_risk':
            for cell in ws[curr_row]: cell.fill = red_fill
        elif 'Total' in df.columns and row['Total'] > 90:
            for cell in ws[curr_row]: cell.fill = green_fill

    wb.save(filename)
    print(f"📄 Excel report saved to: {filename}")

# --- STEP 1: INGESTION & DATA INTEGRITY ---
print(f"📚 Course: {course} | 📄 File: {csv_file}")
try:
    df = pd.read_csv(csv_file)
    # Remove metadata rows unique to Canvas exports
    df = df[df['name'] != 'Points Possible']
    
    # Calculate 'at_risk' status before saving to DB
    if 'Total' in df.columns:
        df["status"] = df["Total"].apply(lambda x: "at_risk" if x < 60 else "passing")
        
except Exception as e:
    print(f"❌ Error loading CSV: {e}")
    sys.exit()

# --- STEP 2: DATABASE PERSISTENCE (ORM Layer) ---
with engine.connect() as conn:
    # This maps our Python Object (DataFrame) to a SQL Data Table
    df.to_sql(course, engine, if_exists="replace", index=False)
    conn.commit()
print("✅ Database Synchronized.")

# --- STEP 3: DYNAMIC SMART INTERFACE ---
print("\n--- 🛠️ DYNAMIC QUERY INTERFACE ---")
print(f"Querying table: {course}")
print("Usage: <SQL Query> [--excel]")
print("Example: SELECT name, Total FROM {course} WHERE Total < 70 --excel")
print("Type 'exit' to quit.")

while True:
    raw_input = input("SQL > ").strip()
    
    if raw_input.lower() == 'exit':
        break
    if not raw_input:
        continue

    # Logic: Only generate excel if requested with the flag
    user_query = raw_input.strip()


    try:
        with engine.connect() as conn:
            if user_query.lower().startswith("select"):
                query_results = pd.read_sql(text(user_query), conn)
                print("\n📊 Results:")
                print(query_results,flush=True)
                execel = input(" Do you want excel:(Y/N): ").lower().strip()

                if execel == "y":
                    report_name = f"{course}_report.xlsx"
                    generate_excel_from_df(query_results, report_name)
                
            else:
                # Handle administr ative commands like DELETE or DROP
                result = conn.execute(text(user_query))
                conn.commit()
                print(f"✅ Command executed successfully.")
                
    except Exception as e:
        print(f"❌ SQL Error: {e}")

print("\n✅ Workflow Complete.")